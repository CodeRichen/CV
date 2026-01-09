#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# 取得腳本所在目錄
script_dir = os.path.dirname(os.path.abspath(__file__))

# 讀取詳細指標CSV
detailed_df = pd.read_csv(os.path.join(script_dir, 'output', 'detailed_metrics_all_classifiers.csv'))
overall_df = pd.read_csv(os.path.join(script_dir, 'output', 'overall_performance.csv'))

# 解析分類器名稱
classifiers = ['KNN (K=7)', 'SVM (RBF)', 'SVM (Linear)', 'LDA', 'Random Forest (n=200)']

# 初始化結果字典
results = {}

for clf in classifiers:
    clf_data = detailed_df[detailed_df['Classifier'] == clf]
    
    # 提取平均指標 (從overall_df)
    overall_row = overall_df[overall_df['Classifier'] == clf].iloc[0]
    
    # 訓練和測試準確率（實際計算）
    test_accuracy = overall_row['Testing_Accuracy']
    train_accuracy = overall_row['Training_Accuracy']
    
    # 平均指標
    avg_precision = overall_row['Avg_Precision']
    avg_recall = overall_row['Avg_Recall']
    avg_f1 = overall_row['Avg_F1-Score']
    avg_dice = overall_row['Avg_Dice']
    avg_iou = overall_row['Avg_IoU']
    
    # 訓練時間
    training_time = overall_row['Training_Time_Seconds']
    
    # 計算每個類別的準確率 (正確率)
    class_accuracy = []
    class_names = []
    for class_name in clf_data['Class'].unique():
        class_data = clf_data[clf_data['Class'] == class_name].iloc[0]
        acc = class_data['Recall']  # 使用Recall作為該類別的準確率
        class_accuracy.append(acc)
        class_names.append(class_name)
    
    # 找出準確率最好和最差的類別（Top 3和Bottom 3）
    accuracy_with_names = list(zip(class_names, class_accuracy))
    accuracy_with_names.sort(key=lambda x: x[1], reverse=True)
    
    top3 = accuracy_with_names[:3]
    bottom3 = accuracy_with_names[-3:]
    
    # 計算標準差
    std_accuracy = np.std(class_accuracy)
    
    # 計算過擬合間隙（訓練準確率 - 測試準確率）
    overfit_gap = train_accuracy - test_accuracy
    
    results[clf] = {
        'Training_Time': training_time,
        'Avg_Precision': avg_precision,
        'Avg_Recall': avg_recall,
        'Avg_F1': avg_f1,
        'Avg_Dice': avg_dice,
        'Avg_IoU': avg_iou,
        'Train_Accuracy': train_accuracy,
        'Test_Accuracy': test_accuracy,
        'Overfit_Gap': overfit_gap,
        'Top3': top3,
        'Bottom3': bottom3,
        'Std_Accuracy': std_accuracy
    }

# 建立主要表格
data = []
for clf in classifiers:
    r = results[clf]
    data.append({
        'Method': clf,
        'Training_Time': f"{r['Training_Time']:.4f}s",
        'Precision': f"{r['Avg_Precision']:.4f}",
        'Recall': f"{r['Avg_Recall']:.4f}",
        'F1-Score': f"{r['Avg_F1']:.4f}",
        'Dice': f"{r['Avg_Dice']:.4f}",
        'IoU': f"{r['Avg_IoU']:.4f}",
        'Training_Accuracy': f"{r['Train_Accuracy']:.4f}",
        'Testing_Accuracy': f"{r['Test_Accuracy']:.4f}",
        'Overfit_Gap': f"{r['Overfit_Gap']:.4f}",
        'Best_Class_1': f"{r['Top3'][0][0]} ({r['Top3'][0][1]:.4f})",
        'Best_Class_2': f"{r['Top3'][1][0]} ({r['Top3'][1][1]:.4f})",
        'Best_Class_3': f"{r['Top3'][2][0]} ({r['Top3'][2][1]:.4f})",
        'Worst_Class_1': f"{r['Bottom3'][0][0]} ({r['Bottom3'][0][1]:.4f})",
        'Worst_Class_2': f"{r['Bottom3'][1][0]} ({r['Bottom3'][1][1]:.4f})",
        'Worst_Class_3': f"{r['Bottom3'][2][0]} ({r['Bottom3'][2][1]:.4f})",
        'Std_Accuracy': f"{r['Std_Accuracy']:.4f}"
    })

df_main = pd.DataFrame(data)

# 建立數值表格（用於計算）
data_numeric = []
for clf in classifiers:
    r = results[clf]
    data_numeric.append({
        'Method': clf,
        'Training_Time': r['Training_Time'],
        'Precision': r['Avg_Precision'],
        'Recall': r['Avg_Recall'],
        'F1-Score': r['Avg_F1'],
        'Dice': r['Avg_Dice'],
        'IoU': r['Avg_IoU'],
        'Training_Accuracy': r['Train_Accuracy'],
        'Testing_Accuracy': r['Test_Accuracy'],
        'Overfit_Gap': r['Overfit_Gap'],
        'Std_Accuracy': r['Std_Accuracy']
    })

df_numeric = pd.DataFrame(data_numeric)

# 建立Excel檔案
excel_path = os.path.join(script_dir, 'Classification_Report.xlsx')
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    # 第一個Sheet：主要報告
    df_main.to_excel(writer, sheet_name='Summary', index=False)
    
    # 第二個Sheet：數值表格
    df_numeric.to_excel(writer, sheet_name='Numeric Data', index=False)
    
    # 第三個Sheet：詳細指標
    detailed_df.to_excel(writer, sheet_name='Detailed Metrics', index=False)

# 格式化Excel檔案
wb = openpyxl.load_workbook(excel_path)

# 格式化Summary Sheet
ws = wb['Summary']
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 標題樣式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)

# 應用標題樣式
for col_num, header in enumerate(df_main.columns, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# 調整列寬
ws.column_dimensions['A'].width = 20
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
    ws.column_dimensions[col].width = 18

# 應用數據行樣式
for row_num in range(2, len(df_main) + 2):
    for col_num in range(1, len(df_main.columns) + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 交替行背景色
        if row_num % 2 == 0:
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

# 格式化Numeric Data Sheet
ws_numeric = wb['Numeric Data']
for col_num, header in enumerate(df_numeric.columns, 1):
    cell = ws_numeric.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

ws_numeric.column_dimensions['A'].width = 20
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
    ws_numeric.column_dimensions[col].width = 16

for row_num in range(2, len(df_numeric) + 2):
    for col_num in range(1, len(df_numeric.columns) + 1):
        cell = ws_numeric.cell(row=row_num, column=col_num)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if col_num > 1:  # 數值列
            cell.number_format = '0.0000'
        if row_num % 2 == 0:
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

# 格式化Detailed Metrics Sheet
ws_detailed = wb['Detailed Metrics']
for col_num, header in enumerate(detailed_df.columns, 1):
    cell = ws_detailed.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for col in range(1, len(detailed_df.columns) + 1):
    ws_detailed.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

for row_num in range(2, len(detailed_df) + 2):
    for col_num in range(1, len(detailed_df.columns) + 1):
        cell = ws_detailed.cell(row=row_num, column=col_num)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if col_num > 1:
            cell.number_format = '0.0000'
        if row_num % 2 == 0:
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

# 保存Excel檔案
wb.save(excel_path)

print(f"✓ Excel報告已生成: {excel_path}")
print("\n=== 報告內容概覽 ===")
print("\nSheet 1: Summary (摘要)")
print(df_main.to_string(index=False))
print("\n\nSheet 2: Numeric Data (數值數據)")
print(df_numeric.to_string(index=False))
print(f"\n\n✓ 報告已保存到: {excel_path}")
print("\n表格包含以下信息:")
print("  - 5個分類方法的性能比較")
print("  - 執行時間、Precision、Recall、F1-Score、Dice、IoU")
print("  - 訓練準確率、測試準確率、過擬合間隙")
print("  - 表現最好/最差的3個圖片類型及其準確率")
print("  - 全部類型準確率的標準差")
print("  - 詳細的每類指標數據")
